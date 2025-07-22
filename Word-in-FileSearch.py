import pandas as pd
import os
import logging
from openpyxl import load_workbook
import re  # Import the regular expression module
from urllib.parse import urlparse # Import for URL parsing

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("excel_search.log"),
        logging.StreamHandler(),
    ],
)


def is_module_installed(module_name: str) -> bool:
    """
    Checks if a Python module is installed.

    Args:
        module_name (str): The name of the module (e.g., 'PIL').

    Returns:
        bool: True if the module is installed, False otherwise.
    """
    try:
        __import__(module_name)
        return True
    except ImportError:
        return False



def install_module(module_name: str, package_name: str = None) -> bool:
    """
    Installs a Python module using pip.

    Args:
        module_name (str): The name of the module (e.g., 'PIL').
        package_name (str, optional): The name of the package to install (if different
            from the module name). Defaults to the module name.

    Returns:
        bool: True if the installation was successful, False otherwise.
    """
    if package_name is None:
        package_name = module_name
    try:
        logging.info(f"Installing module: {module_name} using pip...")
        # Use subprocess.run with check=True to raise an exception on failure
        result = subprocess.run(
            ["pip", "install", package_name],
            check=True,  # Raise an exception for non-zero return codes
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,  # Get output as text
        )
        logging.info(f"Installation successful. Output:\n{result.stdout}")
        return True
    except subprocess.CalledProcessError as e:
        logging.error(
            f"Failed to install module {module_name}. Error:\n{e.stderr}\nOutput:\n{e.stdout}"
        )
        print(f"Error installing {module_name}: e")
        print(f"  Check the error message and ensure pip is installed and working correctly.")
        return False
    except Exception as e:
        logging.error(f"An unexpected error occurred during installation: {e}")
        print(f"An unexpected error occurred: e")
        return False



def search_excel_files(folder_path):
    """
    Searches all Excel files in a given folder for "Level 1" and checks the
    adjacent cell for "Line Manager", even with potential special characters.

    Args:
        folder_path (str): The path to the folder containing the Excel files.

    Returns:
        list: A list of dictionaries, where each dictionary contains the
            file name, sheet name, cell number of "Level Manager", and whether
            "Line Manager" was found in the adjacent cell.
            Returns an empty list if no matching files are found or an error occurs.
    """
    results = []
    try:
        # Check if the path is a URL (SharePoint)
        parsed_url = urlparse(folder_path)
        if parsed_url.scheme in ('http', 'https'):
            print("Error: SharePoint paths are not supported. Please provide a local folder path.")
            logging.error("SharePoint paths are not supported.")
            return []

        # If it is a local path
        if os.path.isdir(folder_path): # Check if folder exists
            for filename in os.listdir(folder_path):
                if filename.endswith((".xlsx", ".xls")):
                    file_path = os.path.join(folder_path, filename)
                    logging.info(f"Processing file: {file_path}")
                    try:
                        workbook = load_workbook(file_path, read_only=True)
                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            logging.info(f"  Processing sheet: {sheet_name}")
                            for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row,
                                                      min_col=sheet.min_column, max_col=sheet.max_column):
                                for cell in row:
                                    if cell.value == "Level 1":
                                        logging.info(f"    Found 'Level 1' at cell {cell.coordinate} in sheet {sheet_name} of file {filename}")
                                        adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                                        # Use regex to find "Line Manager" with any surrounding characters
                                        if adjacent_cell.value is not None and re.search(r".*Line Manager.*", str(adjacent_cell.value), re.IGNORECASE):
                                            logging.info(f"    'Line Manager' found in adjacent cell {adjacent_cell.coordinate}")
                                            results.append(
                                                {
                                                    "file_name": filename,
                                                    "sheet_name": sheet_name,
                                                    "cell_number": cell.coordinate,
                                                    "line_manager_found": True,
                                                }
                                            )
                                        else:
                                            logging.info(f"    'Line Manager' not found in adjacent cell.")
                                            results.append(
                                                {
                                                    "file_name": filename,
                                                    "sheet_name": sheet_name,
                                                    "cell_number": cell.coordinate,
                                                    "line_manager_found": False,
                                                }
                                            )
                        workbook.close()
                    except Exception as e:
                        logging.error(f"Error processing file {file_path}: {e}")
                        print(f"Error processing file {file_path}: {e}")
                        continue
        else:
            print(f"Error: The provided path is not a valid local directory: {folder_path}")
            logging.error(f"Invalid folder path: {folder_path}")
            return []
        return results
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        print(f"An error occurred: {e}")
        return []



def main():
    """
    Main function to call the search function and print the results in a tabular format and save to a text file.
    """
    folder_path = input("Enter the path to the folder containing the Excel files: ")
    results = search_excel_files(folder_path)

    if not results:
        print("No matching files found or an error occurred.")
        logging.info("No matching files found or an error occurred.")
        return

    df = pd.DataFrame(results)
    output_text = df.to_string(index=False)
    print(output_text)

    # Save to a text file in the same folder as the Excel files.
    #  Get the directory from the input folder path
    output_dir = os.path.dirname(folder_path)
    if not output_dir:
        output_dir = "."  # Use current directory if input path is just a folder name
    txt_filename = os.path.join(output_dir, "excel_search_results.txt")
    try:
        with open(txt_filename, "w", encoding="utf-8") as f: #added encoding
            f.write(output_text)
        logging.info(f"Results saved to {txt_filename}")
        print(f"Results saved to {txt_filename}")
    except Exception as e:
        logging.error(f"Error saving results to text file: {e}")
        print(f"Error saving results to text file: {e}")

    # Optionally, save the results to a CSV file
    csv_filename = os.path.join(output_dir, "excel_search_results.csv")
    df.to_csv(csv_filename, index=False)
    logging.info(f"Results saved to {csv_filename}")
    print(f"Results saved to {csv_filename}")



if __name__ == "__main__":
    main()
